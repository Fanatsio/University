from  openpyxl import  *

class Save_Wallpaper():
    def __init__(self):
        pass
    def  data_Save(*args):
        wb=load_workbook("D:\Университет\Pythonlabs\lab_13\Result.xlsx")
        ws = wb.active

        ws['A1'] = "Обои:"
        ws['B1'] = "|"
        ws['C1'] = "Плитка:"
        ws['D1'] = "|"
        ws['E1'] = "Ламинат:"
        ws['F1'] = "|"

        if args[0] != 0:
            ws['A2'] = "Количество:"
            ws['B2'] = args[0]

            ws['A3'] = "Общая цена:"
            ws['B3'].value = args[1]
            wb.save("D:\Университет\Pythonlabs\lab_13\Result.xlsx")

class Save_Tile():
    def __init__(self):
        pass
    def  data_Save(*args):
        wb=load_workbook("D:\Университет\Pythonlabs\lab_13\Result.xlsx")
        ws = wb.active

        ws['A1'] = "Обои:"
        ws['B1'] = "|"
        ws['C1'] = "Плитка:"
        ws['D1'] = "|"
        ws['E1'] = "Ламинат:"
        ws['F1'] = "|"

        if args[0] != 0:
            ws['C2'] = "Количество:"
            ws['D2'] = args[0]

            ws['C3'] = "Общая цена:"
            ws['D3'].value = args[1]
            wb.save("D:\Университет\Pythonlabs\lab_13\Result.xlsx")

class Save_Laminat():
    def __init__(self):
        pass
    def  data_Save(*args):
        wb=load_workbook("D:\Университет\Pythonlabs\lab_13\Result.xlsx")
        ws = wb.active

        ws['A1'] = "Обои:"
        ws['B1'] = "|"
        ws['C1'] = "Плитка:"
        ws['D1'] = "|"
        ws['E1'] = "Ламинат:"
        ws['F1'] = "|"

        if args[0] != 0:
            ws['E2'] = "Количество:"
            ws['F2'] = args[0]
            ws['E3'] = "Общая цена:"
            ws['F3'].value = args[1]
            wb.save("D:\Университет\Pythonlabs\lab_13\Result.xlsx")